import openpyxl
from openpyxl import load_workbook,Workbook

def reading_excel(filename):
    book=load_workbook(filename)

    #active method get the currently active sheet in excel
    sheet=book.active

    #Accessing single value from sheet
    print(sheet['A1'].value)
    cell_obj = sheet['A1': 'B5']

    for cell1, cell2 in cell_obj:
        print(cell1.value,cell2.value)


    print(sheet.cell(row=1,column=2).value)

    #Iterating all values of sheet
    for row in sheet.iter_rows(values_only=True):
        print(row)

    #getting total no of rows and columns
    rows=sheet.max_row
    columns=sheet.max_column
    print(f"Total Rows {rows}")
    print(f"Total Columns {columns}")

    #Modyfying The Data ,After modyfing we need to save it to mmake changes
    sheet['A2'].value="haseeb"
    book.save(filename)


def writing_excel(filename):
    book=load_workbook(filename)
    sheet=book.active

    #Multiple data can be added using append in tuple
    data=(
        ("Ali",88,"B"),
        ("Ali",88,"B")
    )
    for row in data:
        sheet.append(row)
    book.save(filename)

writing_excel('data.xlsx')

reading_excel('data.xlsx')
